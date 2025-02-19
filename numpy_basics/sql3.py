import openpyxl

path = "C:\\FIL Intern\\numpy_basics\\bar_menu.xlsx"

try:
    wb_obj = openpyxl.load_workbook(path)
    sh_obj = wb_obj.active  
except Exception as e:
    print("Error:", e)

rows = sh_obj.max_row
cols = sh_obj.max_column


menu_data = {}

for i in range(1, cols + 1):
    column_name = sh_obj.cell(row=1, column=i).value
    for x in range(2,rows+1):
        menu_data.setdefault(column_name,[]).append(sh_obj.cell(row=x,column=i).value)
    #menu_data[column_name] = [sh_obj.cell(row=x, column=i).value for x in range(2, rows + 1)]

print(menu_data)